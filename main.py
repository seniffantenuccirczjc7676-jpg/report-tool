import io
import re
import datetime
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(page_title="留存周期自动计算与精细填报工具", layout="wide")
st.title("📊 智能报表自动化填报工具")

# -----------------------------------------------------------------------------
# 1. 基础辅助与标准化函数
# -----------------------------------------------------------------------------

def normalize_header(val):
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip()
    if s.startswith("="):
        return ""
    s = s.replace("\u3000", "").replace("\n", "").replace("\r", "").replace("\t", "")
    s = re.sub(r"[\s\u200b\u200c\u200d\ufeff]+", "", s)
    return s

def clean_channel_str(val):
    """超级干净的渠道字符串清洗，统一横杠与各种不可见字符"""
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"[－—–\─]", "-", s)
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"[\s\u3000\u200b\u200c\u200d\ufeff]+", "", s)
    return s.lower()

def find_channel_sheet(wb, raw_channel_name):
    """【渠道 Sheet 名称智能匹配器】"""
    if not raw_channel_name or str(raw_channel_name).strip() == "":
        return None, "渠道名为空"
    
    raw_str = str(raw_channel_name).strip()
    sheet_names = wb.sheetnames

    if raw_str in sheet_names:
        return raw_str, "完全匹配"

    target_clean = clean_channel_str(raw_str)
    for sname in sheet_names:
        if clean_channel_str(sname) == target_clean:
            return sname, "标准化匹配"
            
    return None, "未找到Sheet"

def parse_num(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return 0.0
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_to_date_obj(val):
    """万能日期解析器"""
    if val is None or pd.isna(val):
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, (int, float)):
        try:
            dt = openpyxl.utils.datetime.from_excel(val)
            if isinstance(dt, (datetime.datetime, datetime.date)):
                return dt.date() if isinstance(dt, datetime.datetime) else dt
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return None
    s_date_part = s.split(" ")[0].split("T")[0]
    s_date_part = s_date_part.replace(".", "-").replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    
    parts = s_date_part.split("-")
    if len(parts) == 3:
        try:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime.date(y, m, d)
        except ValueError:
            pass
            
    try:
        pdt = pd.to_datetime(s_date_part, errors="coerce")
        if not pd.isna(pdt):
            return pdt.date()
    except Exception:
        pass
    return None

def build_date_to_row_index(ws, date_col_idx):
    """提取总表中指定列的全部有效日期和对应行号"""
    date_map = {}
    if not date_col_idx:
        return date_map
        
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=date_col_idx).value
        d_obj = parse_to_date_obj(cell_val)
        if d_obj and d_obj not in date_map:
            date_map[d_obj] = r
    return date_map

# -----------------------------------------------------------------------------
# 2. 高效防死锁合并单元格解构与动态列搜索函数
# -----------------------------------------------------------------------------

def get_cell_value_merged(ws, row, col):
    """高效且安全的合并单元格取值逻辑（原生坐标范围快速定位）"""
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for rng in ws.merged_cells.ranges:
                if row >= rng.min_row and row <= rng.max_row and col >= rng.min_col and col <= rng.max_col:
                    return ws.cell(row=rng.min_row, column=rng.min_col).value
        return cell.value
    except Exception:
        return None

def detect_actual_header_rows(ws, date_col_idx=1, max_search=300):
    """自动匹配最可能的表头关键行（全面覆盖深层表头场景，例如163行）"""
    HEADER_KEYWORDS = ["首存", "次日", "留存", "人数", "金额", "ARPPU", "百分比", "3日", "5日", "7日", "15日", "30日"]
    max_r = min(ws.max_row or 300, max_search)
    max_c = min(ws.max_column or 50, 50)

    best_row = 3
    max_matches = 0

    for r in range(1, max_r + 1):
        row_str = ""
        for c in range(1, max_c + 1):
            v = get_cell_value_merged(ws, r, c)
            if v is not None:
                row_str += str(v) + " "

        match_count = sum(1 for kw in HEADER_KEYWORDS if kw in row_str)
        if match_count > max_matches:
            max_matches = match_count
            best_row = r

    return best_row

def find_metric_column(ws, metric_type, module_type, date_col_idx=1):
    """
    【精准按表头名称动态搜索目标列，严格隔离次日与3日、裂变与全盘/直属】
    """
    header_row = detect_actual_header_rows(ws, date_col_idx=date_col_idx)
    col_max = min(ws.max_column or 50, 50)
    best_col = None
    best_match_text = ""
    highest_score = -1

    # 上下扫描3行以补全多级表头的完整上下文
    search_rows = [r for r in [header_row - 1, header_row, header_row + 1] if r > 0]

    for col in range(1, col_max + 1):
        cell_texts = []
        for r in search_rows:
            val = get_cell_value_merged(ws, r, col)
            if val is not None:
                if parse_to_date_obj(val) is not None:
                    continue
                cleaned = normalize_header(val)
                if cleaned and not cleaned.startswith("=") and cleaned not in cell_texts:
                    cell_texts.append(cleaned)
        
        full_path = "/".join(cell_texts)

        if not full_path:
            continue

        # ------------------ 规则 1：模块隔离 ------------------
        if module_type == "全盘人数":
            if any(k in full_path for k in ["直属", "裂变"]) and "全盘" not in full_path:
                continue
        elif module_type == "裂变人数":
            if "裂变" not in full_path:
                continue
        elif module_type in ["直属人数", "直属金额"]:
            if any(k in full_path for k in ["全盘", "裂变"]) and "直属" not in full_path:
                continue

        # ------------------ 规则 2：人数与金额及“率/LTV/均值”排他 ------------------
        if module_type == "直属金额":
            if "金额" not in full_path:
                continue
        elif module_type in ["全盘人数", "直属人数", "裂变人数"]:
            if any(bad_kw in full_path for bad_kw in ["金额", "%", "率", "ltv", "LTV", "均", "ARPU", "arpu"]):
                continue

        # ------------------ 规则 3：周期精细正则匹配 ------------------
        matched_kw = False
        
        if metric_type == 0:  # 首存 / 首充
            if any(k in full_path for k in ["首存", "首充"]):
                matched_kw = True
        elif metric_type == 1:  # 次日 / 1日 / +1
            if "次日" in full_path or re.search(r"(?<!\d)1[日天](?!\d)", full_path) or "+1" in full_path:
                matched_kw = True
        elif metric_type == 2:  # 3日/3天/2日/+2
            if "次日" not in full_path:
                if re.search(r"(?<!\d)[23][日天](?!\d)", full_path) or "+2" in full_path or "+3" in full_path or "30" not in full_path:
                    if any(k in full_path for k in ["3日", "3天", "2日", "2天", "+2", "+3", "d2", "d3", "D2", "D3"]):
                        matched_kw = True
        elif metric_type == 4:  # 5日/5天
            if re.search(r"(?<!\d)5[日天](?!\d)", full_path) or "+5" in full_path:
                matched_kw = True
        elif metric_type == 6:  # 7日/7天
            if re.search(r"(?<!\d)7[日天](?!\d)", full_path) or "+7" in full_path:
                matched_kw = True
        elif metric_type == 14: # 15日/15天
            if re.search(r"(?<!\d)15[日天](?!\d)", full_path) or "+15" in full_path:
                matched_kw = True
        elif metric_type == 29: # 30日/30天
            if re.search(r"(?<!\d)30[日天](?!\d)", full_path) or "+30" in full_path:
                matched_kw = True

        if matched_kw:
            score = 10

            if module_type == "全盘人数" and "全盘" in full_path:
                score += 20
            elif module_type == "裂变人数" and "裂变" in full_path:
                score += 20
            elif "直属" in module_type and "直属" in full_path:
                score += 20

            if module_type == "直属金额" and "金额" in full_path:
                score += 15

            if any(k in full_path for k in ["人数", "留存人数", "首存", "首充"]):
                score += 15

            if score > highest_score:
                highest_score = score
                best_col = col
                best_match_text = full_path

    if best_col is not None:
        return best_col, best_match_text
    return None, "未搜索到目标字段"

# -----------------------------------------------------------------------------
# 3. 核心业务配置与周期常量
# -----------------------------------------------------------------------------

CUMULATIVE_THRESHOLDS = [0, 1, 2, 4, 6, 14, 29]
DIFF_TO_THRESH = {0: 0, 1: 1, 2: 2, 4: 4, 6: 6, 14: 14, 29: 29}
THRESH_NAME_MAP = {0: "0日(首存)", 1: "1日(次日)", 2: "3日", 4: "5日", 6: "7日", 14: "15日", 29: "30日"}

# -----------------------------------------------------------------------------
# Streamlit 界面主体
# -----------------------------------------------------------------------------
col_up1, col_up2 = st.columns(2)
with col_up1:
    source_file = st.file_uploader("1. 上传数据源表格", type=["xlsx", "xls"])
with col_up2:
    master_file = st.file_uploader("2. 上传目标总主表", type=["xlsx", "xls"])

if source_file and master_file:
    try:
        original_master_name = master_file.name
        source_bytes = source_file.read()
        master_bytes = master_file.read()

        excel_mem = pd.ExcelFile(io.BytesIO(source_bytes))
        sheet_names = excel_mem.sheet_names

        wb_master_preview = openpyxl.load_workbook(io.BytesIO(master_bytes), read_only=True)
        master_sheet_names = wb_master_preview.sheetnames

        st.markdown("---")

        # 1. 全盘留存配置
        st.subheader("🔵 1. 【全盘留存人数】通道")
        default_qp_idx = 0
        for idx, sname in enumerate(sheet_names):
            if normalize_header(sname) == "全盘留存":
                default_qp_idx = idx
                break
            elif "全盘" in sname:
                default_qp_idx = idx

        qp_sheet_selected = st.selectbox("选择【全盘留存】源 Sheet", sheet_names, index=default_qp_idx, key="qp_sheet_key")
        qp_df = excel_mem.parse(qp_sheet_selected)
        qp_df.columns = [normalize_header(c) for c in qp_df.columns]
        qp_cols = qp_df.columns.tolist()

        qp_master_sheet_name = "全盘留存" if "全盘留存" in master_sheet_names else master_sheet_names[0]
        ws_master_qp = wb_master_preview[qp_master_sheet_name]

        master_col_options = {}
        for c_idx in range(1, ws_master_qp.max_column + 1):
            val_t = get_cell_value_merged(ws_master_qp, 1, c_idx) or f"第{c_idx}列"
            master_col_options[c_idx] = f"第 {c_idx} 列 ({normalize_header(val_t)})"

        c_qp1, c_qp2, c_qp3 = st.columns(3)
        with c_qp1:
            qp_date_col = st.selectbox("【数据源】日期列", qp_cols, key="qp_d_key")
        with c_qp2:
            qp_master_date_col_idx = st.selectbox(
                "【总表】日期列（请选择包含日期的列）",
                options=list(master_col_options.keys()),
                format_func=lambda x: master_col_options[x],
                index=0,
                key="qp_m_d_key"
            )
        with c_qp3:
            qp_tot_col = st.selectbox("【全盘留存】首充/首存人数列", qp_cols, key="qp_t_key")

        qp_options = ["-- 不填/跳过 --"] + qp_cols
        cq1, cq2, cq3 = st.columns(3)
        with cq1:
            q_m2 = st.selectbox("全盘-次日留存人数", qp_options, key="q_m2")
            q_m3 = st.selectbox("全盘-3日留存人数", qp_options, key="q_m3")
        with cq2:
            q_m5 = st.selectbox("全盘-5日留存人数", qp_options, key="q_m5")
            q_m7 = st.selectbox("全盘-7日留存人数", qp_options, key="q_m7")
        with cq3:
            q_m15 = st.selectbox("全盘-15日留存人数", qp_options, key="q_m15")
            q_m30 = st.selectbox("全盘-30日留存人数", qp_options, key="q_m30")

        qp_cols_map = {
            0: qp_tot_col, 
            1: q_m2 if q_m2 != "-- 不填/跳过 --" else None,
            2: q_m3 if q_m3 != "-- 不填/跳过 --" else None, 
            4: q_m5 if q_m5 != "-- 不填/跳过 --" else None,
            6: q_m7 if q_m7 != "-- 不填/跳过 --" else None, 
            14: q_m15 if q_m15 != "-- 不填/跳过 --" else None,
            29: q_m30 if q_m30 != "-- 不填/跳过 --" else None,
        }

        st.markdown("---")

        # 2. 裂变留存配置
        st.subheader("🟣 2. 【裂变留存人数】通道")
        use_lf_sheet = st.checkbox("处理【裂变留存】？", value=True, key="ulfs")
        lf_sheet_selected, lf_df = None, None
        lf_date_col, lf_tot_col = "时间", "首存"
        lf_cols_map = {}

        if use_lf_sheet:
            default_lf_idx = 0
            for idx, sname in enumerate(sheet_names):
                if "裂变" in sname:
                    default_lf_idx = idx
                    break

            lf_sheet_selected = st.selectbox("选择【裂变留存】源 Sheet", sheet_names, index=default_lf_idx, key="lf_sheet_key")
            lf_df = excel_mem.parse(lf_sheet_selected)
            lf_df.columns = [normalize_header(c) for c in lf_df.columns]
            lf_cols = lf_df.columns.tolist()

            cl1, cl2 = st.columns(2)
            with cl1:
                lf_date_col = st.selectbox("【裂变留存】日期列", lf_cols, key="lf_d_key")
            with cl2:
                lf_tot_col = st.selectbox("【裂变留存】首充/首存人数列", lf_cols, key="lf_t_key")

            lf_options = ["-- 不填/跳过 --"] + lf_cols
            clf1, clf2, clf3 = st.columns(3)
            with clf1:
                l_m2 = st.selectbox("裂变-次日留存人数", lf_options, key="l_m2")
                l_m3 = st.selectbox("裂变-3日留存人数", lf_options, key="l_m3")
            with clf2:
                l_m5 = st.selectbox("裂变-5日留存人数", lf_options, key="l_m5")
                l_m7 = st.selectbox("裂变-7日留存人数", lf_options, key="l_m7")
            with clf3:
                l_m15 = st.selectbox("裂变-15日留存人数", lf_options, key="l_m15")
                l_m30 = st.selectbox("裂变-30日留存人数", lf_options, key="l_m30")

            lf_cols_map = {
                0: lf_tot_col, 
                1: l_m2 if l_m2 != "-- 不填/跳过 --" else None,
                2: l_m3 if l_m3 != "-- 不填/跳过 --" else None, 
                4: l_m5 if l_m5 != "-- 不填/跳过 --" else None,
                6: l_m7 if l_m7 != "-- 不填/跳过 --" else None, 
                14: l_m15 if l_m15 != "-- 不填/跳过 --" else None,
                29: l_m30 if l_m30 != "-- 不填/跳过 --" else None,
            }

        st.markdown("---")

        # 3. 直属留存人数配置
        st.subheader("🟢 3. 【直属留存人数】通道")
        default_zs_idx = 0
        for idx, sname in enumerate(sheet_names):
            if "直属" in sname and "人数" in sname:
                default_zs_idx = idx
                break

        zs_sheet_selected = st.selectbox("选择【直属人数】源 Sheet", sheet_names, index=default_zs_idx, key="zs_sheet_key")
        zs_df = excel_mem.parse(zs_sheet_selected)
        zs_df.columns = [normalize_header(c) for c in zs_df.columns]
        zs_cols = zs_df.columns.tolist()

        cz1, cz2 = st.columns(2)
        with cz1:
            zs_date_col = st.selectbox("【直属人数】日期列", zs_cols, key="zs_d_key")
            zs_tot_col = st.selectbox("【直属人数】首充/首存人数列", zs_cols, key="zs_t_key")
        with cz2:
            has_channel = st.checkbox("包含【主渠道】列（填充直属渠道 Sheet）", value=True, key="zs_hc_key")
            zs_chan_col = st.selectbox("【直属人数】主渠道列", zs_cols, key="zs_c_key") if has_channel else None

        zs_options = ["-- 不填/跳过 --"] + zs_cols
        cz_m1, cz_m2, cz_m3 = st.columns(3)
        with cz_m1:
            z_m2 = st.selectbox("直属-次日留存人数", zs_options, key="z_m2")
            z_m3 = st.selectbox("直属-3日留存人数", zs_options, key="z_m3")
        with cz_m2:
            z_m5 = st.selectbox("直属-5日留存人数", zs_options, key="z_m5")
            z_m7 = st.selectbox("直属-7日留存人数", zs_options, key="z_m7")
        with cz_m3:
            z_m15 = st.selectbox("直属-15日留存人数", zs_options, key="z_m15")
            z_m30 = st.selectbox("直属-30日留存人数", zs_options, key="z_m30")

        zs_cols_map = {
            0: zs_tot_col, 
            1: z_m2 if z_m2 != "-- 不填/跳过 --" else None,
            2: z_m3 if z_m3 != "-- 不填/跳过 --" else None, 
            4: z_m5 if z_m5 != "-- 不填/跳过 --" else None,
            6: z_m7 if z_m7 != "-- 不填/跳过 --" else None, 
            14: z_m15 if z_m15 != "-- 不填/跳过 --" else None,
            29: z_m30 if z_m30 != "-- 不填/跳过 --" else None,
        }

        st.markdown("---")

        # 4. 直属留存金额配置
        st.subheader("💰 4. 【直属留存金额】通道")
        use_amount_sheet = st.checkbox("处理【直属留存金额】？", value=True, key="uas")
        amt_sheet_selected, amt_df = None, None
        amt_date_col, amt_chan_col, amt_val_col = "时间", "主渠道", "金额"

        if use_amount_sheet:
            default_amt_idx = 0
            for idx, sname in enumerate(sheet_names):
                if "直属金额" in sname or "金额" in sname:
                    default_amt_idx = idx
                    break

            amt_sheet_selected = st.selectbox("选择【直属金额】源 Sheet", sheet_names, index=default_amt_idx, key="amt_sheet_key")
            amt_df = excel_mem.parse(amt_sheet_selected)
            amt_df.columns = [normalize_header(c) for c in amt_df.columns]
            amt_cols = amt_df.columns.tolist()

            ca1, ca2, ca3 = st.columns(3)
            with ca1:
                amt_date_col = st.selectbox("金额源-【日期/时间】列", amt_cols, key="amt_d_key")
            with ca2:
                amt_chan_col = st.selectbox("金额源-【主渠道】列", amt_cols, key="amt_c_key")
            with ca3:
                amt_val_col = st.selectbox("金额源-【金额】列", amt_cols, key="amt_v_key")

        st.markdown("---")

        # 🚀 执行按钮
        if st.button("🚀 执行精准动态搜索与填报", type="primary"):
            wb = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=False)
            audit_logs = []
            column_search_logs = []

            target_channel_sheets = [
                s for s in wb.sheetnames 
                if s not in ["全盘留存", "直属留存", "裂变留存", qp_master_sheet_name]
            ]

            # =================================================================
            # A. 处理【全盘留存人数】
            # =================================================================
            qp_df["clean_dt_obj"] = qp_df[qp_date_col].apply(parse_to_date_obj)
            qp_valid = qp_df.dropna(subset=["clean_dt_obj"]).copy()

            ws_qp = wb[qp_master_sheet_name]
            qp_date_idx_map = build_date_to_row_index(ws_qp, qp_master_date_col_idx)

            qp_search_res = {}
            for t in CUMULATIVE_THRESHOLDS:
                c_idx, c_path = find_metric_column(ws_qp, metric_type=t, module_type="全盘人数", date_col_idx=qp_master_date_col_idx)
                qp_search_res[t] = c_idx
                column_search_logs.append({
                    "Sheet名称": qp_master_sheet_name, "模块类型": "全盘人数",
                    "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                    "匹配到的实际表头路径": c_path
                })

            if not qp_valid.empty:
                qp_max_dt = qp_valid["clean_dt_obj"].max()

                for d_obj, grp in qp_valid.groupby("clean_dt_obj"):
                    diff = (qp_max_dt - d_obj).days
                    r_idx = qp_date_idx_map.get(d_obj)

                    if not r_idx:
                        continue

                    for thresh in CUMULATIVE_THRESHOLDS:
                        col_target = qp_search_res.get(thresh)
                        if not col_target:
                            continue

                        if diff >= thresh:
                            src_col = qp_cols_map.get(thresh)
                            if src_col and src_col in grp.columns:
                                val = grp[src_col].apply(parse_num).sum()
                                if val is not None:
                                    ws_qp.cell(row=r_idx, column=col_target).value = val
                                    audit_logs.append({
                                        "模块": "全盘人数", "日期": str(d_obj), "渠道": "全盘汇总",
                                        "目标 Sheet": qp_master_sheet_name, "写入位置": f"第{r_idx}行, 第{col_target}列",
                                        "写入数值": val, "状态": "✅ 写入成功"
                                    })

            # =================================================================
            # B. 处理【裂变留存人数】
            # =================================================================
            if use_lf_sheet and lf_df is not None and not lf_df.empty:
                lf_df["clean_dt_obj"] = lf_df[lf_date_col].apply(parse_to_date_obj)
                lf_valid = lf_df.dropna(subset=["clean_dt_obj"]).copy()

                lf_master_sheet_name = "裂变留存" if "裂变留存" in wb.sheetnames else qp_master_sheet_name
                if lf_master_sheet_name in wb.sheetnames:
                    ws_lf = wb[lf_master_sheet_name]
                    lf_date_idx_map = build_date_to_row_index(ws_lf, qp_master_date_col_idx)

                    lf_search_res = {}
                    for t in CUMULATIVE_THRESHOLDS:
                        c_idx, c_path = find_metric_column(ws_lf, metric_type=t, module_type="裂变人数", date_col_idx=qp_master_date_col_idx)
                        lf_search_res[t] = c_idx
                        column_search_logs.append({
                            "Sheet名称": lf_master_sheet_name, "模块类型": "裂变人数",
                            "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                            "匹配到的实际表头路径": c_path
                        })

                    if not lf_valid.empty:
                        lf_max_dt = lf_valid["clean_dt_obj"].max()

                        for d_obj, grp in lf_valid.groupby("clean_dt_obj"):
                            diff = (lf_max_dt - d_obj).days
                            r_idx = lf_date_idx_map.get(d_obj)

                            if not r_idx:
                                continue

                            for thresh in CUMULATIVE_THRESHOLDS:
                                col_target = lf_search_res.get(thresh)
                                if not col_target:
                                    continue

                                if diff >= thresh:
                                    src_col = lf_cols_map.get(thresh)
                                    if src_col and src_col in grp.columns:
                                        val = grp[src_col].apply(parse_num).sum()
                                        if val is not None:
                                            ws_lf.cell(row=r_idx, column=col_target).value = val
                                            audit_logs.append({
                                                "模块": "裂变人数", "日期": str(d_obj), "渠道": "裂变汇总",
                                                "目标 Sheet": lf_master_sheet_name, "写入位置": f"第{r_idx}行, 第{col_target}列",
                                                "写入数值": val, "状态": "✅ 写入成功"
                                            })

            # =================================================================
            # C. 处理【直属留存人数】
            # =================================================================
            zs_df["clean_dt_obj"] = zs_df[zs_date_col].apply(parse_to_date_obj)
            zs_valid = zs_df.dropna(subset=["clean_dt_obj"]).copy()

            if not zs_valid.empty:
                zs_max_dt = zs_valid["clean_dt_obj"].max()

                # C1. 写入【直属留存】主表
                if "直属留存" in wb.sheetnames:
                    ws_zs = wb["直属留存"]
                    zs_date_idx_map = build_date_to_row_index(ws_zs, qp_master_date_col_idx)

                    zs_main_num_search = {}
                    for t in CUMULATIVE_THRESHOLDS:
                        c_idx, c_path = find_metric_column(ws_zs, metric_type=t, module_type="直属人数", date_col_idx=qp_master_date_col_idx)
                        zs_main_num_search[t] = c_idx
                        column_search_logs.append({
                            "Sheet名称": "直属留存", "模块类型": "直属人数(主表)",
                            "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                            "匹配到的实际表头路径": c_path
                        })

                    for d_obj, grp in zs_valid.groupby("clean_dt_obj"):
                        diff = (zs_max_dt - d_obj).days
                        thresh = DIFF_TO_THRESH.get(diff)
                        if thresh is not None:
                            c = zs_cols_map.get(thresh)
                            if c and c in grp.columns:
                                tot_val = grp[c].apply(parse_num).sum()
                                r_idx = zs_date_idx_map.get(d_obj)
                                col_target = zs_main_num_search.get(thresh)
                                if r_idx and col_target and (tot_val is not None):
                                    ws_zs.cell(row=r_idx, column=col_target).value = tot_val
                                    audit_logs.append({
                                        "模块": "直属人数", "日期": str(d_obj), "渠道": "所有渠道加总",
                                        "目标 Sheet": "直属留存", "写入位置": f"第{r_idx}行, 第{col_target}列",
                                        "写入数值": tot_val, "状态": "✅ 汇总成功"
                                    })

                # C2. 按渠道分别写入对应分 Sheet
                if has_channel and zs_chan_col and zs_chan_col in zs_valid.columns:
                    channel_num_search_cache = {}
                    zs_valid["clean_chan"] = zs_valid[zs_chan_col].apply(clean_channel_str)
                    valid_dates = sorted(zs_valid["clean_dt_obj"].dropna().unique())

                    for d_obj in valid_dates:
                        diff = (zs_max_dt - d_obj).days
                        thresh = DIFF_TO_THRESH.get(diff)
                        
                        if thresh is None:
                            continue

                        c = zs_cols_map.get(thresh)
                        if not c or c not in zs_valid.columns:
                            continue

                        for target_sname in target_channel_sheets:
                            ws_ch = wb[target_sname]
                            target_sname_clean = clean_channel_str(target_sname)

                            if target_sname not in channel_num_search_cache:
                                channel_num_search_cache[target_sname] = {}
                                for t in CUMULATIVE_THRESHOLDS:
                                    c_idx, c_path = find_metric_column(ws_ch, metric_type=t, module_type="直属人数", date_col_idx=qp_master_date_col_idx)
                                    channel_num_search_cache[target_sname][t] = c_idx
                                    column_search_logs.append({
                                        "Sheet名称": target_sname, "模块类型": "直属人数(分渠道Sheet)",
                                        "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                                        "匹配到的实际表头路径": c_path
                                    })

                            ch_date_idx_map = build_date_to_row_index(ws_ch, qp_master_date_col_idx)
                            r_idx = ch_date_idx_map.get(d_obj)

                            if not r_idx:
                                continue

                            col_target = channel_num_search_cache[target_sname].get(thresh)
                            if not col_target:
                                continue

                            matched_rows = zs_valid[
                                (zs_valid["clean_dt_obj"] == d_obj) & 
                                (zs_valid["clean_chan"] == target_sname_clean)
                            ]

                            if not matched_rows.empty:
                                ch_val = matched_rows[c].apply(parse_num).sum()
                                status_str = "✅ 分渠道写入"
                            else:
                                ch_val = 0.0
                                status_str = "🟡 数据源无该渠道，已写入0"

                            ws_ch.cell(row=r_idx, column=col_target).value = ch_val
                            audit_logs.append({
                                "模块": "直属人数", "日期": str(d_obj), "渠道": target_sname,
                                "目标 Sheet": target_sname, "写入位置": f"第{r_idx}行, 第{col_target}列",
                                "写入数值": ch_val, "状态": status_str
                            })

            # =================================================================
            # D. 处理【直属留存金额】（补全完成）
            # =================================================================
            if use_amount_sheet and amt_df is not None and not amt_df.empty:
                amt_df["clean_dt_obj"] = amt_df[amt_date_col].apply(parse_to_date_obj)
                amt_df["amt_num"] = amt_df[amt_val_col].apply(parse_num)
                amt_valid = amt_df.dropna(subset=["clean_dt_obj"]).copy()

                if not amt_valid.empty:
                    amt_max_dt = amt_valid["clean_dt_obj"].max()

                    # D1. 写入【直属留存】主表金额列
                    if "直属留存" in wb.sheetnames:
                        ws_zs = wb["直属留存"]
                        zs_date_idx_map = build_date_to_row_index(ws_zs, qp_master_date_col_idx)

                        zs_main_amt_search = {}
                        for t in CUMULATIVE_THRESHOLDS:
                            c_idx, c_path = find_metric_column(ws_zs, metric_type=t, module_type="直属金额", date_col_idx=qp_master_date_col_idx)
                            zs_main_amt_search[t] = c_idx
                            column_search_logs.append({
                                "Sheet名称": "直属留存", "模块类型": "直属金额(主表)",
                                "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                                "匹配到的实际表头路径": c_path
                            })

                        for d_obj, grp in amt_valid.groupby("clean_dt_obj"):
                            diff = (amt_max_dt - d_obj).days
                            thresh = DIFF_TO_THRESH.get(diff)
                            if thresh is not None:
                                day_tot_amt = grp["amt_num"].sum()
                                r_idx = zs_date_idx_map.get(d_obj)
                                col_target = zs_main_amt_search.get(thresh)
                                if r_idx and col_target and (day_tot_amt is not None):
                                    ws_zs.cell(row=r_idx, column=col_target).value = day_tot_amt
                                    audit_logs.append({
                                        "模块": "直属金额", "日期": str(d_obj), "渠道": "所有渠道加总",
                                        "目标 Sheet": "直属留存", "写入位置": f"第{r_idx}行, 第{col_target}列",
                                        "写入数值": day_tot_amt, "状态": "✅ 汇总成功"
                                    })

                    # D2. 按渠道分别写入各渠道 Sheet
                    channel_amt_search_cache = {}
                    amt_valid["clean_chan"] = amt_valid[amt_chan_col].apply(clean_channel_str)
                    amt_dates = sorted(amt_valid["clean_dt_obj"].dropna().unique())

                    for d_obj in amt_dates:
                        diff = (amt_max_dt - d_obj).days
                        thresh = DIFF_TO_THRESH.get(diff)

                        if thresh is None:
                            continue

                        for target_sname in target_channel_sheets:
                            ws_ch = wb[target_sname]
                            target_sname_clean = clean_channel_str(target_sname)

                            if target_sname not in channel_amt_search_cache:
                                channel_amt_search_cache[target_sname] = {}
                                for t in CUMULATIVE_THRESHOLDS:
                                    c_idx, c_path = find_metric_column(ws_ch, metric_type=t, module_type="直属金额", date_col_idx=qp_master_date_col_idx)
                                    channel_amt_search_cache[target_sname][t] = c_idx
                                    column_search_logs.append({
                                        "Sheet名称": target_sname, "模块类型": "直属金额(分渠道Sheet)",
                                        "指标周期": THRESH_NAME_MAP[t], "匹配列号": f"第{c_idx}列" if c_idx else "❌ 未找到",
                                        "匹配到的实际表头路径": c_path
                                    })

                            ch_date_idx_map = build_date_to_row_index(ws_ch, qp_master_date_col_idx)
                            r_idx = ch_date_idx_map.get(d_obj)

                            if not r_idx:
                                continue

                            col_target = channel_amt_search_cache[target_sname].get(thresh)
                            if not col_target:
                                continue

                            matched_rows = amt_valid[
                                (amt_valid["clean_dt_obj"] == d_obj) & 
                                (amt_valid["clean_chan"] == target_sname_clean)
                            ]

                            if not matched_rows.empty:
                                ch_amt = matched_rows["amt_num"].sum()
                                status_str = "✅ 分渠道写入金额"
                            else:
                                ch_amt = 0.0
                                status_str = "🟡 数据源无该渠道金额，已写入0"

                            ws_ch.cell(row=r_idx, column=col_target).value = ch_amt
                            audit_logs.append({
                                "模块": "直属金额", "日期": str(d_obj), "渠道": target_sname,
                                "目标 Sheet": target_sname, "写入位置": f"第{r_idx}行, 第{col_target}列",
                                "写入数值": ch_amt, "状态": status_str
                            })

            # =================================================================
            # E. 导出结果与日志展示
            # =================================================================
            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)

            st.success("🎉 计算与填报全流程完成！")

            st.download_button(
                label="📥 点击下载填充完成的自动化总报表",
                data=out_buf.getvalue(),
                file_name=f"已自动填报_{original_master_name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.markdown("---")
            t1, t2 = st.tabs(["🔍 列动态搜寻日志", "📋 数据写入明细审计"])

            with t1:
                st.write("##### 自动识别的列索引与路径映射：")
                st.dataframe(pd.DataFrame(column_search_logs), use_container_width=True)

            with t2:
                st.write("##### 具体填报数值明细日志：")
                st.dataframe(pd.DataFrame(audit_logs), use_container_width=True)

    except Exception as e:
        st.error(f"❌ 运行过程中出现异常: {str(e)}")
        st.exception(e)
